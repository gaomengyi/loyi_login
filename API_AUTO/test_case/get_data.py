from turtle import pd

from tool.http_request import *
from openpyxl import load_workbook
from tool.project_path import *


class GetData:
    data = None
    wb = load_workbook(testcase_path)
    sheet = wb['init']
    tel_1 = sheet.cell(2, 1).value
    tel = sheet.cell(3, 1).value
    id = sheet.cell(4, 1).value
    token = None


if __name__ == '__main__':
    g = GetData()
