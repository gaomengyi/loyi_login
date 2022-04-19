# encoding=utf-8
from openpyxl import load_workbook
from test_case.get_data import GetData
from tool.get_config import *
from tool.project_path import *

"""
专门读取excel的值
"""


class GetExcel:
    def __init__(self, filename, sheet):
        self.filename = filename
        self.sheet = sheet

    # 第二种方法
    def get_data(self):
        # tel = "15712932600"
        # tel_1 = "admin"
        get_file = load_workbook(self.filename)  # 打开文件
        get_sheet = get_file[self.sheet]  # 打开sheet页
        rownum = get_sheet.max_row  # 获得最大行值
        tel = getattr(GetData, "tel")
        tel_1 = getattr(GetData, "tel_1")
        test_data = []
        for i in range(2, rownum + 1):
            sub_data = {}
            sub_data['case'] = get_sheet.cell(i, 1).value
            sub_data['modul'] = get_sheet.cell(i, 2).value
            sub_data['title'] = get_sheet.cell(i, 3).value
            sub_data['method'] = get_sheet.cell(i, 4).value
            sub_data['url'] = get_sheet.cell(i, 5).value
            # sub_data['data'] = get_sheet.cell(i,6).value
            if get_sheet.cell(i, 6).value.find("${tel}") != -1:  # 如果在第6列找到${tel}这个值 -1表示没有找到
                # 利用反射：如果在第6列能找到${tel}则进行下面的反射操作
                sub_data['data'] = get_sheet.cell(i, 6).value.replace('${tel}', str(tel))
                print(str(getattr(GetData, 'tel')))
            elif get_sheet.cell(i, 6).value.find("${tel_1") != -1:  # 如果在第6列找到${tel_1}这个值
                sub_data["data"] = get_sheet.cell(i, 6).value.replace('${tel_1}', str(tel_1))
            else:
                sub_data['data'] = get_sheet.cell(i, 6).value
            sub_data['expect'] = get_sheet.cell(i, 7).value
            test_data.append(sub_data)
        print(test_data)
        return test_data

    # 写一个方法：专门存储返回数据
    def data_write(self, i, test_result, result):
        file = load_workbook(self.filename)  # 打开文件
        sheet = file[self.sheet]  # 打开sheet页
        sheet.cell(i, 8).value = test_result  # 定位到i行0弟8列写入test_result值
        sheet.cell(i, 9).value = result  # 定位到i行0弟9列写入result值
        file.save(self.filename)  # 保存文件


if __name__ == '__main__':
    res = GetExcel(testcase_path, 'creat_q')
    res.get_data()
